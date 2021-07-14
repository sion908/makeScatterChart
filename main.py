import openpyxl
import numpy as np
from openpyxl.chart import BarChart, Reference, Series,ScatterChart
from pprint import pprint

def main(wb):
    s = (((-1,1),funma,"元"),((-30,30),funMa,"フーリエ変換"))
    charts = []
    for xrange,formula,name in s:
        X = makeXs(xrange,count=1000)
        datas = [[name+"偶数",(0,2,4)],[name+"奇数",(1,3,5)]]
        print(datas)
        for sheetName,ns in datas:
            x_cols=[]
            for i in ns:
                if i==0 or i==1:
                    ws,row,col=writeExcel(wb,X,formula(X,i),sheetName,str(i))
                else:
                # row ,ws=writeExcel(wb,X,funMa(X,i),"wani",str(i))
                    ws,row,col=writeExcel(wb,None,formula(X,i),sheetName,str(i))
                
                x_cols.append(col)

            makeScatterChar(ws,row,1,x_cols,ns,place="F8",xrange=xrange)
    
#end main

def funMa(X,n):

    return np.sin(n*np.pi/2-X)/(n*np.pi/2-X)  + np.sin(n*np.pi/2+X) / (n*np.pi/2+X)

def funma(X,n):
    #numpy.where(condition[, x, y])
    return np.where(abs(X)<1,np.cos(n*np.pi*X/2),0)

def makeScatterChar(ws,lastrow,x_col,y_cols,titles,place="D16",xrange=[-30,30]):
        #ScatterChartオブジェクト作成
    chart = ScatterChart('marker')

    #xの範囲を設定
    min_row = 2
    max_row = lastrow-1
    x_values = Reference(ws, min_col = x_col, min_row = min_row+1, max_row = max_row-1)
    
    # x_len = (xrange[1]-xrange[0])/1000
    # print(x_len)
    # print(xrange)
    # print(x_values[-1])
    chart.x_axis.scaling.min = xrange[0]
    chart.x_axis.scaling.max = xrange[1]

    for y_col,title in zip(y_cols,titles):
        #yの範囲を設定
        min_col = y_col
        values = Reference(ws, min_col = min_col, min_row = min_row+1, max_row = max_row-1)
        #グラフの追加
        series = Series(values, x_values, title="n="+str(title))
        chart.series.append(series)
    chart.title = "wa"
    ws.add_chart(chart, place)
#enddef makeScatterChar

def makeXs(xrange=[-1,1],count=100):

    x =np.linspace(xrange[0],xrange[1],count+1)
    return x
#end makeXs


def writeExcel(wb,X,Y,sname,dataname):
    found = True
    for name in wb.sheetnames:
        if name == sname:
            ws = wb[name]
            found = False
            col = ws.max_column +1
            break
    if found:
        ws = wb.create_sheet(title=sname)
        col = 1
    
    
    # シートの追加
    # ws4 = wb.create_sheet(title="Sheet4")
    row=1
    # セルの指定
    # c1=ws.cell(row=1,column=1)
    if type(X)==type(None):
        ws.cell(row = row, column = col, value = dataname)
        row +=1
        for y in Y:
            #numpy.float64のままだとValueErrorが出るのでキャスト
            ws.cell(row = row, column = col, value = float(y))
            row+=1
        
    else:
        ws.cell(row = row, column = col, value = "X"+dataname)
        ws.cell(row = row, column = col+1, value = dataname)
        row +=1
        for x,y in zip(X,Y):
            #numpy.init32のままだとValueErrorが出るのでキャスト
            ws.cell(row = row, column = col, value = float(x))
            #numpy.float64のままだとValueErrorが出るのでキャスト
            ws.cell(row = row, column = col + 1, value = float(y))
            row+= 1
        col += 1

    return ws,row,col

#end writeExxcel

def Preprocessing():
    
    wb = openpyxl.Workbook()
    # wb.remove(wb.worksheets)
    sheets =wb.worksheets
    main(wb)
    for sheet in sheets:
        wb.remove(sheet)
    wb.save("wa.xlsx")
    print("complete")
#end Preprocessing

if __name__=="__main__":
    Preprocessing()
    # main(wb)
    # main("wa")
#end ifmain
